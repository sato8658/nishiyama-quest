import csv
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"

def csv_rows(name, header_row=0):
    with (DATA / name).open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    headers = []
    used = {}
    for value in rows[header_row]:
        base = value or "column"
        used[base] = used.get(base, 0) + 1
        headers.append(base if used[base] == 1 else f"{base}_{used[base]}")
    return [dict(zip(headers, row)) for row in rows[header_row + 1:] if any(cell.strip() for cell in row)]

def xlsx_rows(name):
    sheet = load_workbook(DATA / name, data_only=True).active
    headers = [cell.value or f"column_{cell.column}" for cell in sheet[2]]
    rows = []
    for values in sheet.iter_rows(min_row=3, values_only=True):
        if not any(value is not None for value in values):
            continue
        normalized = {}
        for key, value in zip(headers, values):
            normalized[str(key)] = value.isoformat() if hasattr(value, "isoformat") else value
        rows.append(normalized)
    return rows

payload = {
    "publicToilets": csv_rows("public_toilets.csv"),
    "parking": csv_rows("parking.csv"),
    "busStops": csv_rows("bus_stops.csv"),
    "redPandas": csv_rows("red_pandas.csv"),
    "redPandaIndividuals": xlsx_rows("red_panda_individuals.xlsx"),
    "visitorFlow": csv_rows("visitor_flow_2025.csv", 1),
}

(DATA / "open-data.raw.json").write_text(
    json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
    encoding="utf-8",
)
print({key: len(value) for key, value in payload.items()})
